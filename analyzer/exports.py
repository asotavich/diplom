"""
analyzer/exports.py

Generates downloadable report artefacts from completed AnalysisReport rows.
Kept separate from views.py so the format logic is unit-testable and the
view stays focused on HTTP concerns.

Public surface
--------------
``build_report_excel(report) -> openpyxl.Workbook``
    Two-sheet Excel workbook (Summary + Resource Breakdown).

``build_report_pdf(report) -> bytes``
    Single-file PDF document containing the same sections as the Excel
    summary plus a table of the top external hosts (FR-09). Returned as
    bytes so the caller can stream it straight into an HttpResponse.
"""

from __future__ import annotations

from decimal import Decimal
from io import BytesIO
from typing import TYPE_CHECKING

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

if TYPE_CHECKING:
    from .models import AnalysisReport

# ---------------------------------------------------------------------------
# Colour palette (RGB hex, no leading #)
# ---------------------------------------------------------------------------
_DARK_BLUE = "1F4E79"   # title / section headers
_LIGHT_BLUE = "BDD7EE"  # sub-section headers
_LABEL_BG = "EBF3FB"    # label cells (left column)
_WHITE = "FFFFFF"

# Complexity level colours (matches Excel's built-in conditional-format palette)
_GREEN = "C6EFCE"   # low  (< 15)
_YELLOW = "FFEB9C"  # medium (15–39)
_RED = "FFC7CE"     # high (≥ 40)

_LOW_THRESHOLD = Decimal("15")
_HIGH_THRESHOLD = Decimal("40")


# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------

def _fill(hex_color: str) -> PatternFill:
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def _font(bold: bool = False, size: int = 11, color: str = "000000") -> Font:
    return Font(name="Calibri", bold=bold, size=size, color=color)


def _align(horizontal: str = "left", wrap: bool = False) -> Alignment:
    return Alignment(horizontal=horizontal, vertical="center", wrap_text=wrap)


def _border() -> Border:
    thin = Side(style="thin", color="BFBFBF")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _complexity_level(c: Decimal) -> tuple[str, str]:
    """Return (human label, fill hex) for a given complexity value."""
    if c < _LOW_THRESHOLD:
        return "LOW", _GREEN
    if c < _HIGH_THRESHOLD:
        return "MEDIUM", _YELLOW
    return "HIGH", _RED


# ---------------------------------------------------------------------------
# Public entrypoint
# ---------------------------------------------------------------------------

def build_report_excel(report: "AnalysisReport") -> openpyxl.Workbook:
    """
    Build and return a formatted Workbook for the given AnalysisReport.

    The caller is responsible for writing the workbook to a response stream::

        wb = build_report_excel(report)
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        # ... attach buffer.read() to an HttpResponse

    Parameters
    ----------
    report : AnalysisReport
        Must have status == SUCCESS and project / created_by pre-fetched.
    """
    wb = openpyxl.Workbook()
    wb.active.title = "Summary"
    _build_summary_sheet(wb.active, report)

    if report.raw_metadata:
        ws_breakdown = wb.create_sheet(title="Resource Breakdown")
        _build_breakdown_sheet(ws_breakdown, report.raw_metadata)

    return wb


# ---------------------------------------------------------------------------
# Sheet 1: Summary
# ---------------------------------------------------------------------------

def _build_summary_sheet(ws, report: "AnalysisReport") -> None:
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16

    row = 1

    # ---- Title row ----------------------------------------------------------
    ws.merge_cells(f"A{row}:E{row}")
    c = ws.cell(row=row, column=1, value="FEAnalyzer — Architectural Complexity Report")
    c.font = _font(bold=True, size=14, color=_WHITE)
    c.fill = _fill(_DARK_BLUE)
    c.alignment = _align("center")
    ws.row_dimensions[row].height = 30
    row += 2  # one blank spacer row

    # ---- Metadata block -----------------------------------------------------
    scanned = (
        report.scanned_at.strftime("%Y-%m-%d %H:%M:%S UTC")
        if report.scanned_at
        else "—"
    )
    analyst = report.created_by.get_full_name() or report.created_by.username

    for label, value in [
        ("Report ID",  report.pk),
        ("URL",        report.url),
        ("Project",    report.project.name if report.project else "—"),
        ("Scanned at", scanned),
        ("Analyst",    analyst),
    ]:
        lc = ws.cell(row=row, column=1, value=label)
        vc = ws.cell(row=row, column=2, value=value)
        ws.merge_cells(f"B{row}:E{row}")
        lc.font = _font(bold=True)
        lc.fill = _fill(_LABEL_BG)
        lc.alignment = _align("left")
        lc.border = _border()
        vc.alignment = _align("left", wrap=True)
        vc.border = _border()
        ws.row_dimensions[row].height = 18
        row += 1

    row += 1  # blank

    # ---- Component counts section -------------------------------------------
    _section_header(ws, row, "COMPONENT COUNTS")
    ws.row_dimensions[row].height = 22
    row += 1

    for col, header in enumerate(["Component", "Total", "Internal", "External"], start=1):
        c = ws.cell(row=row, column=col, value=header)
        c.font = _font(bold=True)
        c.fill = _fill(_LABEL_BG)
        c.alignment = _align("center")
        c.border = _border()
    row += 1

    meta = report.raw_metadata or {}
    for label, total, cat_key in [
        ("Links",   report.count_links,   "links"),
        ("Styles",  report.count_styles,  "styles"),
        ("Scripts", report.count_scripts, "scripts"),
    ]:
        cat = meta.get(cat_key, {})
        values = [
            label,
            total,
            cat.get("internal", "—"),
            cat.get("external", "—"),
        ]
        for col, val in enumerate(values, start=1):
            c = ws.cell(row=row, column=col, value=val)
            c.border = _border()
            c.alignment = _align("center" if col > 1 else "left")
        row += 1

    row += 1  # blank

    # ---- Weight coefficients section ----------------------------------------
    _section_header(ws, row, "WEIGHT COEFFICIENTS")
    ws.row_dimensions[row].height = 22
    row += 1

    for col, header in enumerate(["W_links", "W_styles", "W_scripts", "Σ W_i"], start=1):
        c = ws.cell(row=row, column=col, value=header)
        c.font = _font(bold=True)
        c.fill = _fill(_LABEL_BG)
        c.alignment = _align("center")
        c.border = _border()
    row += 1

    w_sum = float(report.weight_links + report.weight_styles + report.weight_scripts)
    for col, val in enumerate([
        float(report.weight_links),
        float(report.weight_styles),
        float(report.weight_scripts),
        round(w_sum, 4),
    ], start=1):
        c = ws.cell(row=row, column=col, value=val)
        c.number_format = "0.0000"
        c.alignment = _align("center")
        c.border = _border()
    row += 2  # blank

    # ---- Complexity index section -------------------------------------------
    ws.merge_cells(f"A{row}:E{row}")
    sec = ws.cell(row=row, column=1, value="COMPLEXITY INDEX (C)")
    sec.font = _font(bold=True, size=12, color=_WHITE)
    sec.fill = _fill(_DARK_BLUE)
    sec.alignment = _align("left")
    ws.row_dimensions[row].height = 22
    row += 1

    # Formula explanation
    ws.merge_cells(f"A{row}:E{row}")
    fml = ws.cell(
        row=row, column=1,
        value="C  =  W_links × N_links  +  W_styles × N_styles  +  W_scripts × N_scripts",
    )
    fml.font = _font(size=10, color="595959")
    fml.alignment = _align("left")
    row += 1

    # Numeric result
    ci = report.complexity_index or Decimal("0")
    level_label, level_color = _complexity_level(ci)

    lc = ws.cell(row=row, column=1, value="Complexity Index (C):")
    lc.font = _font(bold=True)
    lc.fill = _fill(_LABEL_BG)
    lc.border = _border()

    vc = ws.cell(row=row, column=2, value=float(ci))
    vc.number_format = "0.0000"
    vc.font = _font(bold=True, size=16)
    vc.fill = _fill(level_color)
    vc.alignment = _align("center")
    vc.border = _border()
    ws.merge_cells(f"B{row}:E{row}")
    ws.row_dimensions[row].height = 28
    row += 1

    # Level label
    lc2 = ws.cell(row=row, column=1, value="Complexity Level:")
    lc2.font = _font(bold=True)
    lc2.fill = _fill(_LABEL_BG)
    lc2.border = _border()

    vc2 = ws.cell(row=row, column=2, value=level_label)
    vc2.font = _font(bold=True, size=13)
    vc2.fill = _fill(level_color)
    vc2.alignment = _align("center")
    vc2.border = _border()
    ws.merge_cells(f"B{row}:E{row}")
    ws.row_dimensions[row].height = 22


def _section_header(ws, row: int, title: str) -> None:
    ws.merge_cells(f"A{row}:E{row}")
    c = ws.cell(row=row, column=1, value=title)
    c.font = _font(bold=True, size=12)
    c.fill = _fill(_LIGHT_BLUE)
    c.alignment = _align("left")


# ---------------------------------------------------------------------------
# Sheet 2: Resource Breakdown
# ---------------------------------------------------------------------------

def _build_breakdown_sheet(ws, metadata: dict) -> None:
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 12

    row = 1

    # Title
    ws.merge_cells(f"A{row}:D{row}")
    t = ws.cell(row=row, column=1, value="Internal / External Resource Breakdown")
    t.font = _font(bold=True, size=13, color=_WHITE)
    t.fill = _fill(_DARK_BLUE)
    t.alignment = _align("center")
    ws.row_dimensions[row].height = 26
    row += 2

    # Summary table
    for col, h in enumerate(["Category", "Total", "Internal", "External"], start=1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = _font(bold=True)
        c.fill = _fill(_LABEL_BG)
        c.alignment = _align("center")
        c.border = _border()
    row += 1

    for cat in ("links", "styles", "scripts"):
        info = metadata.get(cat, {})
        for col, val in enumerate(
            [cat.capitalize(), info.get("total", 0), info.get("internal", 0), info.get("external", 0)],
            start=1,
        ):
            c = ws.cell(row=row, column=col, value=val)
            c.border = _border()
            c.alignment = _align("center" if col > 1 else "left")
        row += 1

    row += 1

    # Collect and sort external hosts across all categories
    all_hosts: list[tuple[str, str, int]] = []
    for cat in ("links", "styles", "scripts"):
        for entry in metadata.get(cat, {}).get("top_external_hosts", []):
            all_hosts.append((cat.capitalize(), entry["host"], entry["count"]))
    all_hosts.sort(key=lambda x: -x[2])

    if not all_hosts:
        return

    ws.merge_cells(f"A{row}:D{row}")
    hdr = ws.cell(row=row, column=1, value="Top External Hosts")
    hdr.font = _font(bold=True, size=12)
    hdr.fill = _fill(_LIGHT_BLUE)
    hdr.alignment = _align("left")
    ws.row_dimensions[row].height = 20
    row += 1

    for col, h in enumerate(["Category", "Host", "Count", ""], start=1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = _font(bold=True)
        c.fill = _fill(_LABEL_BG)
        c.alignment = _align("center")
        c.border = _border()
    row += 1

    for cat, host, count in all_hosts[:20]:  # cap at 20 rows
        ws.cell(row=row, column=1, value=cat).border = _border()
        ws.cell(row=row, column=2, value=host).border = _border()
        ws.cell(row=row, column=3, value=count).border = _border()
        ws.cell(row=row, column=1).alignment = _align("left")
        ws.cell(row=row, column=2).alignment = _align("left")
        ws.cell(row=row, column=3).alignment = _align("center")
        row += 1


# ---------------------------------------------------------------------------
# PDF — single-page formatted report (FR-09)
# ---------------------------------------------------------------------------

# reportlab uses HTML hex colours rather than Excel fills.
_PDF_DARK_BLUE = colors.HexColor("#1F4E79")
_PDF_LIGHT_BLUE = colors.HexColor("#BDD7EE")
_PDF_LABEL_BG = colors.HexColor("#EBF3FB")
_PDF_GREEN = colors.HexColor("#C6EFCE")
_PDF_YELLOW = colors.HexColor("#FFEB9C")
_PDF_RED = colors.HexColor("#FFC7CE")
_PDF_BORDER = colors.HexColor("#BFBFBF")


def _pdf_level_color(c: Decimal) -> colors.Color:
    if c < _LOW_THRESHOLD:
        return _PDF_GREEN
    if c < _HIGH_THRESHOLD:
        return _PDF_YELLOW
    return _PDF_RED


def build_report_pdf(report: "AnalysisReport") -> bytes:
    """
    Render ``report`` as a printable PDF document and return the bytes.

    The layout mirrors the Excel summary (metadata, counts, weights,
    complexity, top hosts) so users get the same artefact in the format
    that's most convenient for them. reportlab is pure-Python — no
    external binaries — so this works identically on Linux/Windows/Docker.
    """
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=18 * mm,
        rightMargin=18 * mm,
        topMargin=18 * mm,
        bottomMargin=18 * mm,
        title=f"FEAnalyzer report #{report.pk}",
        author="FEAnalyzer",
    )

    styles = getSampleStyleSheet()
    h1 = ParagraphStyle(
        "h1",
        parent=styles["Heading1"],
        textColor=colors.white,
        fontSize=16,
        leading=20,
        backColor=_PDF_DARK_BLUE,
        borderPadding=(8, 8, 8, 8),
        alignment=1,
        spaceAfter=14,
    )
    h2 = ParagraphStyle(
        "h2",
        parent=styles["Heading2"],
        fontSize=12,
        textColor=_PDF_DARK_BLUE,
        spaceBefore=10,
        spaceAfter=6,
    )
    body = styles["BodyText"]
    body_small = ParagraphStyle("body_small", parent=body, fontSize=9, leading=11)

    story = []

    # ---- Title ------------------------------------------------------------
    story.append(Paragraph("FEAnalyzer — Architectural Complexity Report", h1))

    # ---- Metadata ---------------------------------------------------------
    scanned = (
        report.scanned_at.strftime("%Y-%m-%d %H:%M:%S UTC")
        if report.scanned_at else "—"
    )
    analyst = report.created_by.get_full_name() or report.created_by.username
    source_label = (
        report.url
        or (
            report.uploaded_file.name.rsplit("/", 1)[-1]
            if report.uploaded_file else "—"
        )
    )
    source_kind = "URL" if report.source_type == "URL" else "Uploaded HTML file"

    metadata_rows = [
        ["Report ID", str(report.pk)],
        ["Source", f"{source_kind} — {source_label}"],
        ["Project", report.project.name if report.project else "—"],
        ["Scanned at", scanned],
        ["Analyst", analyst],
    ]
    metadata_table = Table(metadata_rows, colWidths=[42 * mm, 130 * mm])
    metadata_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), _PDF_LABEL_BG),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.4, _PDF_BORDER),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(metadata_table)
    story.append(Spacer(1, 12))

    # ---- Component counts -------------------------------------------------
    story.append(Paragraph("Component counts", h2))
    meta = report.raw_metadata or {}
    counts_rows = [["Component", "Total", "Internal", "External"]]
    for label, total, key in (
        ("Links", report.count_links, "links"),
        ("Stylesheets", report.count_styles, "styles"),
        ("Scripts", report.count_scripts, "scripts"),
    ):
        cat = meta.get(key, {}) or {}
        counts_rows.append([
            label,
            str(total),
            str(cat.get("internal", "—")),
            str(cat.get("external", "—")),
        ])
    counts_table = Table(counts_rows, colWidths=[55 * mm, 35 * mm, 40 * mm, 42 * mm])
    counts_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), _PDF_LIGHT_BLUE),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (1, 0), (-1, -1), "CENTER"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.4, _PDF_BORDER),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(counts_table)

    # ---- Weights ----------------------------------------------------------
    story.append(Paragraph("Weight coefficients", h2))
    w_sum = float(report.weight_links + report.weight_styles + report.weight_scripts)
    weights_rows = [
        ["W_links", "W_styles", "W_scripts", "Σ Wᵢ"],
        [
            f"{float(report.weight_links):.4f}",
            f"{float(report.weight_styles):.4f}",
            f"{float(report.weight_scripts):.4f}",
            f"{w_sum:.4f}",
        ],
    ]
    weights_table = Table(weights_rows, colWidths=[43 * mm] * 4)
    weights_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), _PDF_LIGHT_BLUE),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.4, _PDF_BORDER),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(weights_table)

    # ---- Complexity index -------------------------------------------------
    story.append(Paragraph("Complexity index (C)", h2))
    story.append(Paragraph(
        "C = W_links × N_links + W_styles × N_styles + W_scripts × N_scripts",
        body_small,
    ))
    ci = report.complexity_index or Decimal("0")
    level_label, _ = _complexity_level(ci)
    fill_color = _pdf_level_color(ci)
    ci_rows = [
        ["Complexity Index (C)", f"{float(ci):.4f}"],
        ["Complexity Level", level_label],
    ]
    ci_table = Table(ci_rows, colWidths=[60 * mm, 112 * mm])
    ci_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), _PDF_LABEL_BG),
        ("BACKGROUND", (1, 0), (1, -1), fill_color),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
        ("ALIGN", (1, 0), (1, -1), "CENTER"),
        ("FONTSIZE", (0, 0), (-1, -1), 11),
        ("GRID", (0, 0), (-1, -1), 0.4, _PDF_BORDER),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(ci_table)

    # ---- Top external hosts ----------------------------------------------
    all_hosts: list[tuple[str, str, int]] = []
    for cat in ("links", "styles", "scripts"):
        for entry in (meta.get(cat, {}) or {}).get("top_external_hosts", []) or []:
            all_hosts.append((cat.capitalize(), str(entry.get("host", "")), int(entry.get("count", 0))))
    all_hosts.sort(key=lambda x: -x[2])

    if all_hosts:
        story.append(Paragraph("Top external hosts", h2))
        host_rows = [["Category", "Host", "Count"]]
        for cat, host, count in all_hosts[:15]:
            host_rows.append([cat, host, str(count)])
        host_table = Table(host_rows, colWidths=[35 * mm, 110 * mm, 27 * mm])
        host_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), _PDF_LIGHT_BLUE),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("ALIGN", (2, 0), (2, -1), "CENTER"),
            ("GRID", (0, 0), (-1, -1), 0.4, _PDF_BORDER),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        story.append(host_table)

    doc.build(story)
    return buffer.getvalue()
